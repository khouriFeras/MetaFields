#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
json_to_xlsx.py
- Convert JSON product data with meta fields to Excel format
"""

import json
import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment


def convert_json_to_xlsx(input_file: str, output_file: str = None) -> None:
    """Convert JSON product data to Excel format."""
    
    # Load JSON data
    with open(input_file, 'r', encoding='utf-8') as f:
        products = json.load(f)
    
    print(f"📥 Loaded {len(products)} products from {input_file}")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Dog Food Products with Meta Fields"
    
    # Define headers
    headers = [
        "Product ID",
        "Title", 
        "Handle",
        "Product Type",
        "Vendor",
        "Status",
        "Brand Name",
        "Product Type (Meta)",
        "Weight (kg)",
        "Age Group",
        "Target Weight (kg)",
        "Nutritional Benefits",
        "Ingredients", 
        "Special Features"
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write product data
    for row, product in enumerate(products, 2):
        metafields = product.get('metafields', {})
        
        # Product basic info
        ws.cell(row=row, column=1, value=product.get('id', ''))
        ws.cell(row=row, column=2, value=product.get('title', ''))
        ws.cell(row=row, column=3, value=product.get('handle', ''))
        ws.cell(row=row, column=4, value=product.get('productType', ''))
        ws.cell(row=row, column=5, value=product.get('vendor', ''))
        ws.cell(row=row, column=6, value=product.get('status', ''))
        
        # Meta fields
        ws.cell(row=row, column=7, value=metafields.get('brand_name', ''))
        ws.cell(row=row, column=8, value=metafields.get('product_type', ''))
        ws.cell(row=row, column=9, value=metafields.get('weight_kg', ''))
        ws.cell(row=row, column=10, value=metafields.get('age_group', ''))
        ws.cell(row=row, column=11, value=metafields.get('target_weight_kg', ''))
        ws.cell(row=row, column=12, value=metafields.get('nutritional_benefits', ''))
        ws.cell(row=row, column=13, value=metafields.get('ingredients', ''))
        ws.cell(row=row, column=14, value=metafields.get('special_features', ''))
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        for row in range(1, len(products) + 2):
            cell_value = ws[f"{column_letter}{row}"].value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        # Set column width (min 10, max 50)
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save file
    if output_file is None:
        output_file = input_file.replace('.json', '_with_meta_fields.xlsx')
    
    wb.save(output_file)
    
    print(f"✅ Excel file created successfully!")
    print(f"📁 File: {output_file}")
    print(f"📊 Products: {len(products)}")
    print(f"📋 Columns: {len(headers)}")
    
    # Show sample data
    print(f"\n📋 Sample data:")
    if products:
        sample_product = products[0]
        metafields = sample_product.get('metafields', {})
        print(f"  - Title: {sample_product.get('title', '')[:50]}...")
        print(f"  - Brand: {metafields.get('brand_name', 'N/A')}")
        print(f"  - Type: {metafields.get('product_type', 'N/A')}")
        print(f"  - Weight: {metafields.get('weight_kg', 'N/A')} kg")
        print(f"  - Age Group: {metafields.get('age_group', 'N/A')}")


def main():
    parser = argparse.ArgumentParser(description="Convert JSON product data to Excel")
    parser.add_argument('--input-file', required=True, help='Input JSON file')
    parser.add_argument('--output-file', help='Output Excel file')
    
    args = parser.parse_args()
    
    convert_json_to_xlsx(args.input_file, args.output_file)


if __name__ == "__main__":
    main()
