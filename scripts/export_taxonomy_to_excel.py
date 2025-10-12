#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Export Shopify Taxonomy to Excel
Creates a readable Excel file with all categories and their metafields.
"""
import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Fix Windows console encoding
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except:
        pass


def load_taxonomy(taxonomy_file: str = "data/shopify_taxonomy_full.json"):
    """Load taxonomy from JSON file."""
    print(f"📂 Loading taxonomy from: {taxonomy_file}")
    with open(taxonomy_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    print(f"  ✅ Loaded successfully")
    return data


def create_categories_sheet(wb: Workbook, categories: list):
    """Create sheet with all categories."""
    ws = wb.create_sheet("Categories", 0)
    
    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Headers
    headers = ["Category ID", "Full Name", "Short Name", "Level", "Has Metafields", "# of Metafields"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sort categories by full name
    sorted_categories = sorted(categories, key=lambda x: x.get('fullName', ''))
    
    # Data
    row_idx = 2
    for cat in sorted_categories:
        ws.cell(row=row_idx, column=1, value=cat.get('id', ''))
        ws.cell(row=row_idx, column=2, value=cat.get('fullName', ''))
        ws.cell(row=row_idx, column=3, value=cat.get('name', ''))
        ws.cell(row=row_idx, column=4, value=cat.get('level', 0))
        
        # Check if has metafields
        has_metafields = len(cat.get('attributes', [])) > 0
        ws.cell(row=row_idx, column=5, value="Yes" if has_metafields else "No")
        ws.cell(row=row_idx, column=6, value=len(cat.get('attributes', [])))
        
        row_idx += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 40  # ID
    ws.column_dimensions['B'].width = 70  # Full Name
    ws.column_dimensions['C'].width = 30  # Short Name
    ws.column_dimensions['D'].width = 10  # Level
    ws.column_dimensions['E'].width = 15  # Has Metafields
    ws.column_dimensions['F'].width = 15  # # of Metafields
    
    # Freeze first row
    ws.freeze_panes = "A2"
    
    print(f"  ✅ Created Categories sheet with {row_idx-2} categories")


def create_categories_with_metafields_sheet(wb: Workbook, categories: list):
    """Create sheet with only categories that have metafields."""
    ws = wb.create_sheet("Categories with Metafields", 1)
    
    # Header style
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Headers
    headers = ["Full Category Name", "Short Name", "# of Metafields", "Metafield Names"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Filter and sort categories with metafields
    categories_with_mf = [cat for cat in categories if len(cat.get('attributes', [])) > 0]
    sorted_categories = sorted(categories_with_mf, key=lambda x: x.get('fullName', ''))
    
    # Data
    row_idx = 2
    for cat in sorted_categories:
        ws.cell(row=row_idx, column=1, value=cat.get('fullName', ''))
        ws.cell(row=row_idx, column=2, value=cat.get('name', ''))
        ws.cell(row=row_idx, column=3, value=len(cat.get('attributes', [])))
        
        # List metafield names
        metafield_names = [attr.get('name', '') for attr in cat.get('attributes', [])]
        ws.cell(row=row_idx, column=4, value=", ".join(metafield_names))
        
        row_idx += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 70  # Full Name
    ws.column_dimensions['B'].width = 30  # Short Name
    ws.column_dimensions['C'].width = 15  # # of Metafields
    ws.column_dimensions['D'].width = 80  # Metafield Names
    
    # Freeze first row
    ws.freeze_panes = "A2"
    
    print(f"  ✅ Created Categories with Metafields sheet with {row_idx-2} categories")


def create_metafields_sheet(wb: Workbook, categories: list):
    """Create sheet with all metafields details."""
    ws = wb.create_sheet("Metafields Details", 2)
    
    # Header style
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000", size=11)
    
    # Headers
    headers = ["Category", "Metafield Name", "Key", "Type", "Description", "# of Values", "Sample Values"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Collect all metafields with their categories
    metafields_data = []
    for cat in categories:
        for attr in cat.get('attributes', []):
            metafields_data.append({
                'category': cat.get('fullName', ''),
                'name': attr.get('name', ''),
                'handle': attr.get('handle', ''),
                'description': attr.get('description', ''),
                'id': attr.get('id', '')
            })
    
    # Sort by category then metafield name
    sorted_metafields = sorted(metafields_data, key=lambda x: (x['category'], x['name']))
    
    # Load metafield definitions to get types and values
    # We'll need to match by handle from the processed taxonomy
    
    # Data
    row_idx = 2
    for mf in sorted_metafields:
        ws.cell(row=row_idx, column=1, value=mf['category'])
        ws.cell(row=row_idx, column=2, value=mf['name'])
        ws.cell(row=row_idx, column=3, value=mf['handle'])
        ws.cell(row=row_idx, column=4, value="")  # Will be filled from mapping
        ws.cell(row=row_idx, column=5, value=mf['description'])
        ws.cell(row=row_idx, column=6, value="")  # Will be filled if has values
        ws.cell(row=row_idx, column=7, value="")  # Sample values
        
        row_idx += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 60  # Category
    ws.column_dimensions['B'].width = 30  # Name
    ws.column_dimensions['C'].width = 30  # Key
    ws.column_dimensions['D'].width = 25  # Type
    ws.column_dimensions['E'].width = 50  # Description
    ws.column_dimensions['F'].width = 12  # # of Values
    ws.column_dimensions['G'].width = 50  # Sample Values
    
    # Freeze first row
    ws.freeze_panes = "A2"
    
    print(f"  ✅ Created Metafields Details sheet with {row_idx-2} metafields")


def create_categories_with_types_sheet(wb: Workbook, categories_with_metafields: list):
    """Create sheet with categories and their metafields with proper types."""
    ws = wb.create_sheet("Categories + Metafields", 3)
    
    # Header style
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Headers
    headers = ["Category", "Metafield Name", "Key", "Type", "Has Values", "Description"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Collect all data
    row_idx = 2
    for cat in sorted(categories_with_metafields, key=lambda x: x.get('fullName', '')):
        category_name = cat.get('fullName', '')
        
        for mf in cat.get('metafields', []):
            ws.cell(row=row_idx, column=1, value=category_name)
            ws.cell(row=row_idx, column=2, value=mf.get('name', ''))
            ws.cell(row=row_idx, column=3, value=mf.get('key', ''))
            ws.cell(row=row_idx, column=4, value=mf.get('type', ''))
            
            # Check if has predefined values
            values = mf.get('values', [])
            has_values = "Yes" if values and len(values) > 0 else "No"
            ws.cell(row=row_idx, column=5, value=f"{has_values} ({len(values)} options)" if values else "No")
            
            ws.cell(row=row_idx, column=6, value=mf.get('description', ''))
            
            row_idx += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 60  # Category
    ws.column_dimensions['B'].width = 30  # Name
    ws.column_dimensions['C'].width = 30  # Key
    ws.column_dimensions['D'].width = 30  # Type
    ws.column_dimensions['E'].width = 20  # Has Values
    ws.column_dimensions['F'].width = 50  # Description
    
    # Freeze first row
    ws.freeze_panes = "A2"
    
    print(f"  ✅ Created Categories + Metafields sheet with {row_idx-2} entries")


def create_summary_sheet(wb: Workbook, taxonomy_data: dict):
    """Create summary sheet with statistics."""
    ws = wb.create_sheet("Summary", 0)
    
    # Title
    ws['A1'] = "Shopify Product Taxonomy Overview"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    
    # Statistics
    stats = taxonomy_data.get('statistics', {})
    
    row = 3
    stats_data = [
        ("Total Categories", stats.get('total_categories', 0)),
        ("Leaf Categories (Most Specific)", stats.get('leaf_categories', 0)),
        ("Categories with Metafields", stats.get('categories_with_metafields', 0)),
        ("", ""),
        ("Taxonomy Source", "github.com/Shopify/product-taxonomy"),
        ("Data Format", "Open Source (MIT License)"),
        ("Last Updated", "Latest from GitHub"),
    ]
    
    for label, value in stats_data:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1
    
    # Instructions
    row += 2
    ws.cell(row=row, column=1, value="Sheet Guide:").font = Font(bold=True, size=12)
    row += 1
    
    instructions = [
        ("Categories", "All 21,000+ categories (filterable)"),
        ("Categories with Metafields", "8,486 categories that have metafields"),
        ("Categories + Metafields", "Detailed view with types and values"),
        ("Metafields Details", "All metafields with descriptions"),
    ]
    
    for sheet_name, description in instructions:
        ws.cell(row=row, column=1, value=sheet_name).font = Font(bold=True)
        ws.cell(row=row, column=2, value=description)
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 50
    
    print(f"  ✅ Created Summary sheet")


def export_taxonomy_to_excel(taxonomy_file: str = "data/shopify_taxonomy_full.json",
                             output_file: str = "shopify_taxonomy_reference.xlsx"):
    """Export taxonomy to Excel file."""
    print("📊 Exporting Shopify Taxonomy to Excel\n")
    
    # Load taxonomy
    taxonomy_data = load_taxonomy(taxonomy_file)
    
    print("\n📝 Creating Excel sheets...")
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create sheets
    create_summary_sheet(wb, taxonomy_data)
    create_categories_sheet(wb, taxonomy_data.get('all_categories', []))
    create_categories_with_metafields_sheet(wb, taxonomy_data.get('all_categories', []))
    create_categories_with_types_sheet(wb, taxonomy_data.get('categories_with_metafields', []))
    create_metafields_sheet(wb, taxonomy_data.get('all_categories', []))
    
    # Save
    print(f"\n💾 Saving to: {output_file}")
    wb.save(output_file)
    print(f"  ✅ Saved successfully!")
    
    # Summary
    print("\n" + "="*60)
    print("📊 EXCEL FILE CREATED")
    print("="*60)
    print(f"File: {output_file}")
    print("\nSheets:")
    print("  1. Summary - Statistics and overview")
    print("  2. Categories - All 21,000+ categories")
    print("  3. Categories with Metafields - 8,486 categories")
    print("  4. Categories + Metafields - Detailed with types")
    print("  5. Metafields Details - All metafield info")
    print("="*60)
    print("\n✅ Open the file to browse Shopify's taxonomy!")


def main():
    """Main function."""
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Export Shopify taxonomy to Excel for easy browsing"
    )
    parser.add_argument(
        '--taxonomy',
        default='data/shopify_taxonomy_full.json',
        help='Path to taxonomy JSON file (default: data/shopify_taxonomy_full.json)'
    )
    parser.add_argument(
        '--output',
        default='shopify_taxonomy_reference.xlsx',
        help='Output Excel file name (default: shopify_taxonomy_reference.xlsx)'
    )
    
    args = parser.parse_args()
    
    # Check if taxonomy file exists
    if not Path(args.taxonomy).exists():
        print(f"❌ Taxonomy file not found: {args.taxonomy}")
        print("\nPlease run first:")
        print("  python scripts/fetch_shopify_taxonomy.py")
        return
    
    # Export
    export_taxonomy_to_excel(args.taxonomy, args.output)


if __name__ == "__main__":
    main()

