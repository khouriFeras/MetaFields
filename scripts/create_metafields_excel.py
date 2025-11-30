#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Create Excel Report for Category Metafields
Generates a comprehensive Excel file with products and their filled metafields.
"""
import json
import os
import sys
import re
import yaml
from pathlib import Path
from typing import Dict, List, Any, Optional
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Fix Windows console encoding
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except:
        pass


def load_json(file_path: str) -> Any:
    """Load JSON file."""
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)


# Global cache for taxonomy data
_taxonomy_cache: Optional[Dict] = None


def load_taxonomy_data() -> Dict:
    """
    Load Shopify taxonomy data (values and attributes) for normalization.
    Returns a mapping structure for value normalization.
    """
    global _taxonomy_cache
    if _taxonomy_cache is not None:
        return _taxonomy_cache
    
    values_file = Path("data/shopify_values.yml")
    attributes_file = Path("data/shopify_attributes.yml")
    
    if not values_file.exists() or not attributes_file.exists():
        # Return empty cache if files don't exist
        _taxonomy_cache = {"values": {}, "attributes": {}}
        return _taxonomy_cache
    
    # Load values
    with open(values_file, 'r', encoding='utf-8') as f:
        values_data = yaml.safe_load(f) or []
    
    # Build comprehensive value lookup
    # Maps: handle -> name, friendly_id -> name, name variations -> name
    value_map = {}
    for value_item in values_data:
        handle = value_item.get('handle', '')
        name = value_item.get('name', '')
        friendly_id = value_item.get('friendly_id', '')
        
        if name:
            canonical_name = name  # Use the official name as canonical
            
            # Map handle to canonical name (e.g., "color__red" -> "Red")
            if handle:
                value_map[handle] = canonical_name
                value_map[handle.lower()] = canonical_name
                # Also map with underscores and hyphens normalized
                handle_normalized = handle.replace('_', '-').replace('-', '_')
                value_map[handle_normalized] = canonical_name
                value_map[handle_normalized.lower()] = canonical_name
            
            # Map friendly_id to canonical name (e.g., "color__red" -> "Red")
            if friendly_id:
                value_map[friendly_id] = canonical_name
                value_map[friendly_id.lower()] = canonical_name
                # Also map with underscores and hyphens normalized
                friendly_normalized = friendly_id.replace('_', '-').replace('-', '_')
                value_map[friendly_normalized] = canonical_name
                value_map[friendly_normalized.lower()] = canonical_name
            
            # Map name variations to canonical name
            value_map[name.lower()] = canonical_name
            value_map[name] = canonical_name
            
            # Extract value part from handle/friendly_id for partial matching
            # e.g., "color__red" -> also map "red" -> "Red"
            if "__" in handle or "__" in friendly_id:
                source = handle if "__" in handle else friendly_id
                if "__" in source:
                    value_part = source.split("__")[-1]
                    # Normalize value part
                    value_part_clean = value_part.replace("-", " ").replace("_", " ").strip()
                    if value_part_clean and len(value_part_clean) >= 2:
                        value_map[value_part_clean.lower()] = canonical_name
                        value_map[value_part_clean] = canonical_name
    
    # Load attributes to get attribute-to-value mappings
    with open(attributes_file, 'r', encoding='utf-8') as f:
        attributes_data = yaml.safe_load(f) or {}
    
    attribute_map = {}
    if 'base_attributes' in attributes_data:
        for attr in attributes_data['base_attributes']:
            attr_handle = attr.get('handle', '').replace('_', '-')
            if attr_handle:
                attribute_map[attr_handle] = {
                    'name': attr.get('name', ''),
                    'values': attr.get('values', [])
                }
    
    _taxonomy_cache = {
        "values": value_map,
        "attributes": attribute_map
    }
    
    return _taxonomy_cache


def normalize_metafield_value(value: str, metafield: Dict) -> str:
    """
    Normalize metafield values using Shopify taxonomy data.
    Works for all product types by using the actual taxonomy definitions.
    
    Uses the metafield key to look up values in the correct attribute context.
    
    Examples:
    - "hdr-format__hdr10" -> looks up in taxonomy -> "HDR 10"
    - "color__red" -> looks up in taxonomy -> "Red"
    - "احمر" -> Arabic color -> "Red" (if in taxonomy)
    """
    if not value or not isinstance(value, str):
        return str(value) if value else ""
    
    original_value = value.strip()
    if not original_value:
        return ""
    
    # Load taxonomy data
    taxonomy = load_taxonomy_data()
    value_map = taxonomy.get("values", {})
    attribute_map = taxonomy.get("attributes", {})
    
    # Get metafield key to narrow down search (e.g., "color", "hdr-format")
    metafield_key = metafield.get("key", "").replace("_", "-")
    
    # Try to find canonical name from taxonomy
    value_lower = original_value.lower()
    
    # 1. Check if it's already a handle (e.g., "color__red", "hdr-format__hdr10")
    if "__" in original_value or ("-" in original_value and metafield_key):
        # Try exact handle match
        if original_value in value_map:
            return value_map[original_value]
        if value_lower in value_map:
            return value_map[value_lower]
        
        # Extract value part from handle (e.g., "color__red" -> "red")
        if "__" in original_value:
            parts = original_value.split("__")
            if len(parts) >= 2:
                value_part = parts[-1].replace("-", " ").replace("_", " ")
                # Try to find in taxonomy, prioritizing matches for this attribute
                for key, canonical_name in value_map.items():
                    # Check if this value belongs to the current attribute
                    if metafield_key and metafield_key.replace("-", "__") in key:
                        if value_part.lower() in key.lower() or key.lower() in value_part.lower():
                            if len(value_part) >= 3:
                                return canonical_name
                    # Fallback: general match
                    elif value_part.lower() in key.lower() or key.lower() in value_part.lower():
                        if len(value_part) >= 3:
                            return canonical_name
    
    # 2. Check if it's a friendly_id format
    if original_value in value_map:
        return value_map[original_value]
    if value_lower in value_map:
        return value_map[value_lower]
    
    # 3. Use attribute context to narrow search
    if metafield_key and metafield_key in attribute_map:
        attr_info = attribute_map[metafield_key]
        # Get list of valid values for this attribute
        valid_value_ids = attr_info.get("values", [])
        # Try to match against values for this specific attribute
        for value_id in valid_value_ids:
            if value_id in value_map:
                # Check if our value matches this taxonomy value
                canonical_name = value_map[value_id]
                canonical_lower = canonical_name.lower()
                if value_lower == canonical_lower or value_lower in canonical_lower or canonical_lower in value_lower:
                    return canonical_name
                # Check handle format
                if value_id.lower() == value_lower or value_lower in value_id.lower():
                    return canonical_name
    
    # 4. Try partial matching across all taxonomy values
    normalized_input = re.sub(r'[_\-\s]+', ' ', original_value.lower()).strip()
    for key, canonical_name in value_map.items():
        key_normalized = re.sub(r'[_\-\s]+', ' ', key.lower()).strip()
        # Check if normalized input matches or is contained in key
        if normalized_input == key_normalized:
            return canonical_name
        # Check if one contains the other (for partial matches)
        if len(normalized_input) >= 3 and len(key_normalized) >= 3:
            if normalized_input in key_normalized or key_normalized in normalized_input:
                # Prefer exact matches or longer matches
                if abs(len(normalized_input) - len(key_normalized)) <= 2:
                    return canonical_name
    
    # 5. Arabic to English mapping (common colors and terms)
    arabic_to_english = {
        "أحمر": "Red", "احمر": "Red",
        "أزرق": "Blue", "ازرق": "Blue",
        "أخضر": "Green", "اخضر": "Green",
        "أصفر": "Yellow", "اصفر": "Yellow",
        "أسود": "Black", "اسود": "Black",
        "أبيض": "White", "ابيض": "White",
        "رمادي": "Gray", "بني": "Brown",
        "برتقالي": "Orange", "وردي": "Pink",
        "بنفسجي": "Purple", "ذهبي": "Gold",
        "فضي": "Silver",
    }
    
    if original_value in arabic_to_english:
        english_value = arabic_to_english[original_value]
        # Check if English value exists in taxonomy
        if english_value.lower() in value_map:
            return value_map[english_value.lower()]
        return english_value
    
    # 6. If not found in taxonomy, do basic normalization
    # Replace hyphens/underscores with spaces
    normalized = original_value.replace("_", " ").replace("-", " ")
    normalized = " ".join(normalized.split())
    
    # Add spacing for acronyms (e.g., "HDR10" -> "HDR 10")
    if len(normalized) > 1:
        normalized = re.sub(r'([A-Za-z])(\d)', r'\1 \2', normalized)
        normalized = re.sub(r'(\d)([A-Za-z])', r'\1 \2', normalized)
        normalized = " ".join(normalized.split())
    
    # Capitalize if it's a common word (single word, lowercase)
    if " " not in normalized and normalized.islower() and len(normalized) > 2:
        normalized = normalized.capitalize()
    
    return normalized


def create_summary_sheet(wb: Workbook, products: List[Dict], mapping: Dict) -> None:
    """Create summary sheet with overview information."""
    ws = wb.create_sheet("Summary", 0)
    
    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Title
    ws['A1'] = "Category Metafields Analysis"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    
    # Category Information
    row = 3
    ws[f'A{row}'] = "CATEGORY INFORMATION"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    category_info = [
        ("Tag:", mapping.get("tag", "N/A")),
        ("Shopify Category:", mapping["category"]["fullName"]),
        ("Category ID:", mapping["category"]["id"]),
        ("Confidence:", mapping["category"]["confidence"].upper()),
        ("Reasoning:", mapping["category"]["reasoning"]),
    ]
    
    for label, value in category_info:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    # Products Statistics
    row += 1
    ws[f'A{row}'] = "PRODUCTS STATISTICS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    total_products = len(products)
    products_with_metafields = sum(1 for p in products if p.get("category_metafields"))
    
    products_stats = [
        ("Total Products:", total_products),
        ("Products with Metafields:", products_with_metafields),
        ("Coverage:", f"{(products_with_metafields/total_products*100):.1f}%"),
    ]
    
    for label, value in products_stats:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    # Top Vendors
    row += 1
    ws[f'A{row}'] = "TOP VENDORS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    vendors = [p.get('vendor', 'Unknown') for p in products]
    top_vendors = Counter(vendors).most_common(10)
    
    for vendor, count in top_vendors:
        ws[f'A{row}'] = vendor
        ws[f'B{row}'] = count
        row += 1
    
    # Metafields Statistics
    row += 1
    ws[f'A{row}'] = "METAFIELDS STATISTICS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    ws[f'A{row}'] = "Metafield"
    ws[f'B{row}'] = "Filled Count"
    ws[f'C{row}'] = "Coverage %"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    row += 1
    metafield_stats = {}
    for mf in mapping["metafields"]:
        key = mf["key"]
        filled_count = sum(
            1 for p in products 
            if p.get("category_metafields", {}).get(key) is not None
        )
        metafield_stats[mf["name"]] = {
            "filled": filled_count,
            "percentage": (filled_count / total_products * 100) if total_products > 0 else 0
        }
    
    for metafield_name, stats in sorted(metafield_stats.items(), key=lambda x: x[1]["filled"], reverse=True):
        ws[f'A{row}'] = metafield_name
        ws[f'B{row}'] = stats["filled"]
        ws[f'C{row}'] = f"{stats['percentage']:.1f}%"
        row += 1
    
    # Auto-adjust column widths
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 40


def create_products_sheet(wb: Workbook, products: List[Dict], mapping: Dict) -> None:
    """Create products sheet with all products and their metafields."""
    ws = wb.create_sheet("Products", 1)
    
    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define columns
    base_columns = [
        "Handle",
        "Title",
        "Product Type",
        "Vendor",
        "Status"
    ]
    
    # Add metafield columns - use Arabic name from metafield definition
    metafield_columns = [mf.get('name', f"Metafield: shopify.{mf['key']} [{mf['type']}]") for mf in mapping["metafields"]]
    all_columns = base_columns + metafield_columns
    
    # Write headers
    for col_idx, header in enumerate(all_columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Write product data
    for row_idx, product in enumerate(products, 2):
        # Base data
        ws.cell(row=row_idx, column=1, value=product.get('handle', ''))
        ws.cell(row=row_idx, column=2, value=product.get('title', ''))
        ws.cell(row=row_idx, column=3, value=product.get('productType', ''))
        ws.cell(row=row_idx, column=4, value=product.get('vendor', ''))
        ws.cell(row=row_idx, column=5, value=product.get('status', ''))
        
        # Metafield data
        category_metafields = product.get('category_metafields', {})
        for col_idx, mf in enumerate(mapping["metafields"], len(base_columns) + 1):
            value = category_metafields.get(mf["key"])
            
            # Format value based on type
            if value is None:
                display_value = ""
            elif isinstance(value, list):
                # Normalize each value in the list
                normalized_values = [normalize_metafield_value(str(v), mf) for v in value if v]
                display_value = ", ".join(normalized_values)
            elif isinstance(value, dict):
                display_value = json.dumps(value)
            else:
                display_value = normalize_metafield_value(str(value), mf)
            
            cell = ws.cell(row=row_idx, column=col_idx, value=display_value)
            
            # Highlight empty metafields
            if not display_value:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Auto-adjust column widths
    for col_idx in range(1, len(all_columns) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        for row_idx in range(1, len(products) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Freeze first row
    ws.freeze_panes = "A2"


def create_metafields_sheet(wb: Workbook, mapping: Dict) -> None:
    """Create metafields definition sheet."""
    ws = wb.create_sheet("Metafield Definitions", 2)
    
    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ["Name", "Key", "Namespace", "Type", "Description"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write metafield definitions
    for row_idx, mf in enumerate(mapping["metafields"], 2):
        ws.cell(row=row_idx, column=1, value=mf["name"])
        ws.cell(row=row_idx, column=2, value=mf["key"])
        ws.cell(row=row_idx, column=3, value=mf["namespace"])
        ws.cell(row=row_idx, column=4, value=mf["type"])
        ws.cell(row=row_idx, column=5, value=mf.get("description", ""))
    
    # Auto-adjust column widths
    for col_idx in range(1, 6):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        for row_idx in range(1, len(mapping["metafields"]) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)


def create_excel_report(products: List[Dict], mapping: Dict, output_file: str) -> None:
    """Create complete Excel report."""
    print(f"\n Creating Excel report...")
    
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create sheets
    print("   Creating Summary sheet...")
    create_summary_sheet(wb, products, mapping)
    
    print("   Creating Products sheet...")
    create_products_sheet(wb, products, mapping)
    
    print("   Creating Metafield Definitions sheet...")
    create_metafields_sheet(wb, mapping)
    
    # Save workbook
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    
    print(f" Excel report saved: {output_file}")


def main():
    """Main function."""
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Create Excel report for category metafields",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python scripts/create_metafields_excel.py \\
    --products exports/tag_water-pump/products_with_metafields.json \\
    --mapping exports/tag_water-pump/tag_water-pump_category_mapping.json \\
    --output exports/tag_water-pump/water-pump_metafields_final.xlsx
        """
    )
    
    parser.add_argument('--products', required=True, help='Path to products JSON file with metafields')
    parser.add_argument('--mapping', required=True, help='Path to category mapping JSON file')
    parser.add_argument('--output', required=True, help='Output Excel file path')
    
    args = parser.parse_args()
    
    # Load data
    print(f" Loading products from: {args.products}")
    products = load_json(args.products)
    print(f"Loaded {len(products)} products")
    
    print(f"\nLoading category mapping from: {args.mapping}")
    mapping = load_json(args.mapping)
    print(f" Category: {mapping['category']['fullName']}")
    print(f"Metafields: {len(mapping['metafields'])}")
    
    # Create Excel report
    create_excel_report(products, mapping, args.output)
    
    print("\n Excel report creation complete!")
    print(f"\n Review the file: {args.output}")
    print("   - Summary: Overview and statistics")
    print("   - Products: All products with filled metafields")
    print("   - Metafield Definitions: Field specifications")


if __name__ == "__main__":
    main()

