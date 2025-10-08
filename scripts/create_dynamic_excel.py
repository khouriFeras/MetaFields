#!/usr/bin/env python3
"""
Create Excel output for dynamic product analysis results.
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def create_dynamic_analysis_excel(json_file: str, excel_file: str = None) -> str:
    """Create Excel output for dynamic product analysis."""
    try:
        # Load data
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        products = data.get('products', [])
        meta_field_definitions = data.get('meta_field_definitions', {})
        analysis = data.get('product_analysis', {})
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        create_summary_sheet(wb, analysis, meta_field_definitions)
        create_products_sheet(wb, products, meta_field_definitions)
        create_meta_fields_sheet(wb, meta_field_definitions)
        
        # Save file
        if excel_file is None:
            excel_file = str(Path(json_file).parent / f"{Path(json_file).stem}.xlsx")
        
        wb.save(excel_file)
        print(f"Excel file saved: {excel_file}")
        
        return excel_file
        
    except Exception as e:
        print(f"WARNING: Error creating Excel: {e}")
        return None


def create_summary_sheet(wb: Workbook, analysis: dict, meta_field_definitions: dict):
    """Create summary sheet with analysis results."""
    ws = wb.create_sheet("Summary", 0)
    
    # Headers
    headers = [
        "Analysis Type", "Value", "Count", "Percentage"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    row = 2
    
    # Product Analysis Summary
    ws.cell(row=row, column=1, value="Detected Category").font = Font(bold=True)
    ws.cell(row=row, column=2, value=analysis.get('detected_category', 'Unknown'))
    row += 1
    
    ws.cell(row=row, column=1, value="Total Products").font = Font(bold=True)
    ws.cell(row=row, column=2, value=analysis.get('total_products', 0))
    row += 1
    
    ws.cell(row=row, column=1, value="Unique Tags").font = Font(bold=True)
    ws.cell(row=row, column=2, value=analysis.get('unique_tags', 0))
    row += 1
    
    ws.cell(row=row, column=1, value="Unique Vendors").font = Font(bold=True)
    ws.cell(row=row, column=2, value=analysis.get('unique_vendors', 0))
    row += 2
    
    # Top Tags
    ws.cell(row=row, column=1, value="TOP TAGS").font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    row += 1
    
    top_tags = analysis.get('top_tags', {})
    total_products = analysis.get('total_products', 1)
    
    for tag, count in list(top_tags.items())[:10]:
        ws.cell(row=row, column=1, value="Tag")
        ws.cell(row=row, column=2, value=tag)
        ws.cell(row=row, column=3, value=count)
        ws.cell(row=row, column=4, value=f"{(count/total_products)*100:.1f}%")
        row += 1
    
    row += 1
    
    # Top Vendors
    ws.cell(row=row, column=1, value="TOP VENDORS").font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    row += 1
    
    top_vendors = analysis.get('top_vendors', {})
    
    for vendor, count in list(top_vendors.items())[:10]:
        ws.cell(row=row, column=1, value="Vendor")
        ws.cell(row=row, column=2, value=vendor)
        ws.cell(row=row, column=3, value=count)
        ws.cell(row=row, column=4, value=f"{(count/total_products)*100:.1f}%")
        row += 1
    
    row += 1
    
    # Top Product Types
    ws.cell(row=row, column=1, value="TOP PRODUCT TYPES").font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    row += 1
    
    top_product_types = analysis.get('top_product_types', {})
    
    for product_type, count in list(top_product_types.items())[:10]:
        ws.cell(row=row, column=1, value="Product Type")
        ws.cell(row=row, column=2, value=product_type)
        ws.cell(row=row, column=3, value=count)
        ws.cell(row=row, column=4, value=f"{(count/total_products)*100:.1f}%")
        row += 1
    
    row += 1
    
    # Meta Fields Summary
    ws.cell(row=row, column=1, value="DISCOVERED META FIELDS").font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    row += 1
    
    for field_key, field_def in meta_field_definitions.items():
        ws.cell(row=row, column=1, value="Meta Field")
        ws.cell(row=row, column=2, value=field_def.get('name', field_key))
        ws.cell(row=row, column=3, value=field_def.get('type', 'N/A'))
        ws.cell(row=row, column=4, value="Multi-select" if field_def.get('multi_select', False) else "Single-select")
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_products_sheet(wb: Workbook, products: list, meta_field_definitions: dict):
    """Create products sheet with all product data."""
    ws = wb.create_sheet("Products", 1)
    
    # Headers
    # Dynamic headers based on collection type
    headers = ['ID', 'Title', 'Vendor', 'Product Type', 'Tags']
    
    # Add dynamic fields based on discovered meta fields
    dynamic_headers = []
    
    # Map discovered fields to display names
    field_mapping = {
        'brand': 'Brand/Vendor',
        'brand_name': 'Brand/Vendor',
        'brand_vendor': 'Brand/Vendor',
        'product_type': 'Product Type',
        'flavor': 'Flavor',
        'age_range': 'Age/Target',
        'age_target': 'Age/Target',
        'target_audience': 'Target Audience',
        'special_features': 'Features',
        'features': 'Features',
        'key_features': 'Key Features',
        'package_size': 'Size/Weight',
        'size_weight': 'Size/Weight',
        'size_dimensions': 'Size/Weight',
        'price_range': 'Price Range',
        'weight': 'Size/Weight',
        'size': 'Size/Weight',
        'material': 'Material',
        'color': 'Color',
        'origin_country': 'Origin Country',
        'toy_type': 'Toy Type',
        'age_suitability': 'Age Suitability',
        'special_attributes': 'Special Attributes',
        'tool_type': 'Tool Type',
        'power_type': 'Power Type',
        'connectivity': 'Connectivity',
        'compatibility': 'Compatibility'
    }
    
    # Add fields that exist in the data
    for field_key, field_def in meta_field_definitions.items():
        display_name = field_mapping.get(field_key, field_def.get('name', field_key))
        if display_name not in dynamic_headers:
            dynamic_headers.append(display_name)
    
    # Add standard fields
    dynamic_headers.extend(['Price Range', 'Original Price', 'Status'])
    
    headers.extend(dynamic_headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row, product in enumerate(products, 2):
        metafields = product.get('metafields', {})
        
        ws.cell(row=row, column=1, value=product.get('id', ''))
        ws.cell(row=row, column=2, value=product.get('title', ''))
        ws.cell(row=row, column=3, value=product.get('vendor', ''))
        ws.cell(row=row, column=4, value=product.get('productType', ''))
        ws.cell(row=row, column=5, value=', '.join(product.get('tags', [])))
        
        # Dynamic data population based on discovered fields
        col = 6  # Start after basic fields
        
        # Create reverse mapping from display name to field key
        # Since multiple keys can map to same display name, we need to check which keys actually exist
        reverse_mapping = {}
        for k, v in field_mapping.items():
            if k in meta_field_definitions and v not in reverse_mapping:
                reverse_mapping[v] = k
        
        # Populate dynamic fields
        for header in dynamic_headers:
            if header == 'Price Range':
                # Calculate price range from pricing data (prices are in fils, divide by 1000 for JOD)
                # Note: 1 JOD = 1000 fils
                pricing = product.get('pricing', {})
                min_price = pricing.get('min_price', 0)
                if min_price:
                    min_price_jod = min_price / 1000  # Convert fils to JOD (1 JOD = 1000 fils)
                    if min_price_jod < 10:
                        price_range = "Under 10 JOD"
                    elif min_price_jod < 50:
                        price_range = "10-50 JOD"
                    elif min_price_jod < 100:
                        price_range = "50-100 JOD"
                    elif min_price_jod < 200:
                        price_range = "100-200 JOD"
                    else:
                        price_range = "200+ JOD"
                else:
                    price_range = "Not Specified"
                ws.cell(row=row, column=col, value=price_range)
            elif header == 'Original Price':
                # Format price properly (convert from fils to JOD)
                # Note: 1 JOD = 1000 fils
                pricing = product.get('pricing', {})
                min_price = pricing.get('min_price', 0)
                max_price = pricing.get('max_price', 0)
                currency = pricing.get('currency', 'JOD')
                if min_price and max_price:
                    min_jod = min_price / 1000  # Convert fils to JOD
                    max_jod = max_price / 1000  # Convert fils to JOD
                    if min_jod == max_jod:
                        price_display = f"{currency} {min_jod:.2f}"
                    else:
                        price_display = f"{currency} {min_jod:.2f} - {max_jod:.2f}"
                else:
                    price_display = "Price not available"
                ws.cell(row=row, column=col, value=price_display)
            elif header == 'Status':
                ws.cell(row=row, column=col, value=product.get('status', ''))
            else:
                # Find the field key for this header
                field_key = None
                
                # Method 1: Check reverse mapping
                if header in reverse_mapping:
                    field_key = reverse_mapping[header]
                
                # Method 2: Try to find by comparing with meta field definitions (exact match)
                if not field_key:
                    for k, field_def in meta_field_definitions.items():
                        field_name = field_def.get('name', '')
                        if header == field_name:
                            field_key = k
                            break
                
                # Method 3: Partial match - check if header is in field name or vice versa
                if not field_key:
                    for k, field_def in meta_field_definitions.items():
                        field_name = field_def.get('name', '')
                        # Split by parentheses to get English part
                        header_base = header.split('(')[0].strip() if '(' in header else header
                        field_base = field_name.split('(')[0].strip() if '(' in field_name else field_name
                        if header_base == field_base or k in header.lower().replace('/', '_').replace(' ', '_'):
                            field_key = k
                            break
                
                if field_key:
                    # For size_weight field ONLY, prefer standardized weight categories
                    # For "size" field (dimensions), prefer original values
                    # For other fields, prefer original values
                    if field_key == 'size_weight':
                        # Use STANDARDIZED category for weight ranges
                        value = metafields.get(field_key, "")
                        # Fallback to original if standardized is "Not Specified"
                        if value == "Not Specified" or not value:
                            value = metafields.get(f"{field_key}_original", "")
                    else:
                        # For other fields, prefer original values
                        value = metafields.get(f"{field_key}_original", "")
                        # Fallback to standardized if original is empty
                        if not value or value in ["", "null", "None"]:
                            standardized = metafields.get(field_key, "")
                            if standardized and standardized != "Not Specified":
                                value = standardized
                    
                    # Final cleanup
                    if value in ["Not Specified", "null", "None", ""]:
                        value = ""
                    
                    ws.cell(row=row, column=col, value=value)
                else:
                    ws.cell(row=row, column=col, value='')
            col += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_meta_fields_sheet(wb: Workbook, meta_field_definitions: dict):
    """Create meta fields definition sheet."""
    ws = wb.create_sheet("Meta Fields", 2)
    
    # Headers
    headers = [
        'Field Key', 'Name', 'Type', 'Multi-Select', 'Description', 
        'Category', 'Searchable', 'Filterable', 'Comparable', 'Arabic Keywords'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row, (field_key, field_def) in enumerate(meta_field_definitions.items(), 2):
        ws.cell(row=row, column=1, value=field_key)
        ws.cell(row=row, column=2, value=field_def.get('name', ''))
        ws.cell(row=row, column=3, value=field_def.get('type', ''))
        ws.cell(row=row, column=4, value="Yes" if field_def.get('multi_select', False) else "No")
        ws.cell(row=row, column=5, value=field_def.get('description', ''))
        ws.cell(row=row, column=6, value=field_def.get('category', ''))
        ws.cell(row=row, column=7, value="Yes" if field_def.get('searchable', False) else "No")
        ws.cell(row=row, column=8, value="Yes" if field_def.get('filterable', False) else "No")
        ws.cell(row=row, column=9, value="Yes" if field_def.get('comparable', False) else "No")
        ws.cell(row=row, column=10, value=', '.join(field_def.get('arabic_keywords', [])))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def main():
    """Main function."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Create Excel for Dynamic Analysis')
    parser.add_argument('json_file', help='Input JSON file path')
    parser.add_argument('-o', '--output', help='Output Excel file path')
    
    args = parser.parse_args()
    
    try:
        # Create Excel file
        excel_output = create_dynamic_analysis_excel(args.json_file, args.output)
        
        if excel_output:
            print(f"\n Success! Excel file created: {excel_output}")
        else:
            print(" Failed to create Excel file")
            return 1
            
    except Exception as e:
        print(f"Error: {e}")
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main())
